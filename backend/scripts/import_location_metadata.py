"""
Import location metadata from Excel files and icon images.

Reads data/*.xlsx files, maps columns to Location model fields,
and uploads icon images from data/location-icon/{slug}.png.

Usage:
    cd backend
    python scripts/import_location_metadata.py
"""

import os
import sys
import json
import shutil
import uuid

# Ensure the backend directory is on the path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from slugify import slugify
from app import create_app
from app.extensions import db
from app.listing.models import Location, LocationCategory, Listing, SeoAliasSlug

import openpyxl

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
ICON_DIR = os.path.join(DATA_DIR, "location-icon")

# ---------------------------------------------------------------------------
# Column indices (0-based) for the standardised 135-column Excel format
# ---------------------------------------------------------------------------
COL = {
    "name": 1,
    "slug": 2,
    "set": 0,
    "lat": 6,
    "lng": 7,
    # SEO
    "seo_title": 10,
    "seo_description": 11,
    "seo_keywords": 12,
    # Hero / stats
    "population": 22,
    "hero_stat1_value": 29,
    "hero_stat1_label": 30,
    "hero_stat2_value": 31,
    "hero_stat2_label": 32,
    "hero_stat3_value": 33,
    "hero_stat3_label": 34,
    "hero_stat4_value": 35,
    "hero_stat4_label": 36,
    # About
    "about_p1": 25,
    "about_p2": 26,
    # Places (3 × 4 cols)
    "place1_emoji": 37, "place1_tag": 38, "place1_name": 39, "place1_desc": 40,
    "place2_emoji": 41, "place2_tag": 42, "place2_name": 43, "place2_desc": 44,
    "place3_emoji": 45, "place3_tag": 46, "place3_name": 47, "place3_desc": 48,
    # Food
    "food_tags": 50,
    "food1_name": 51, "food1_desc": 52,
    "food2_name": 53, "food2_desc": 54,
    "food3_name": 55, "food3_desc": 56,
    "food4_name": 57, "food4_desc": 58,
    # Sub-areas (3 × 3 cols)
    "sub1_name": 59, "sub1_tag": 60, "sub1_desc": 61,
    "sub2_name": 62, "sub2_tag": 63, "sub2_desc": 64,
    "sub3_name": 65, "sub3_tag": 66, "sub3_desc": 67,
    # Property (3 × 3 cols + rental)
    "prop1_type": 68, "prop1_price": 69, "prop1_subarea": 70,
    "prop2_type": 71, "prop2_price": 72, "prop2_subarea": 73,
    "prop3_type": 74, "prop3_price": 75, "prop3_subarea": 76,
    "rental_range": 77,
    # Schools (4 × 2 cols)
    "school1_name": 78, "school1_board": 79,
    "school2_name": 80, "school2_board": 81,
    "school3_name": 82, "school3_board": 83,
    "school4_name": 84, "school4_board": 85,
    # Hospitals (3 × 2 cols)
    "hospital1_name": 86, "hospital1_type": 87,
    "hospital2_name": 88, "hospital2_type": 89,
    "hospital3_name": 90, "hospital3_type": 91,
    # Commute
    "railway_station": 92,
    "railway_line": 93,
    "time_to_churchgate": 94,
    "metro": 95,
    "highway": 96,
    "bus_routes": 97,
    # Civic
    "municipal_ward": 98,
    "municipal_body": 99,
    "assembly_constituency": 100,
    "lok_sabha_constituency": 101,
    # Economy / Employers (3 × 2 cols)
    "employer1_name": 102, "employer1_sector": 103,
    "employer2_name": 104, "employer2_sector": 105,
    "employer3_name": 106, "employer3_sector": 107,
    # FAQ (5 × 2 cols)
    "faq1_q": 108, "faq1_a": 109,
    "faq2_q": 110, "faq2_a": 111,
    "faq3_q": 112, "faq3_a": 113,
    "faq4_q": 114, "faq4_a": 115,
    "faq5_q": 116, "faq5_a": 117,
    # Nearby (6 × 2 cols)
    "nearby1_slug": 119,
    "nearby2_slug": 121,
    "nearby3_slug": 123,
    "nearby4_slug": 125,
    "nearby5_slug": 127,
    "nearby6_slug": 129,
    # Prose / best_services
    "prose_h2": 130,
    "prose_p1": 131,
    "prose_p2": 132,
    "prose_p3": 133,
    "prose_about": 134,
}

# Map the Excel "Set" column value → zone slug
ZONE_MAP = {
    "West Mumbai": "west-mumbai",
    "South Mumbai": "south-mumbai",
    "Central Mumbai": "central-mumbai",
    "North Mumbai": "north-mumbai",
}

# Display names for zone locations
ZONE_DISPLAY = {
    "west-mumbai": "West Mumbai",
    "south-mumbai": "South Mumbai",
    "central-mumbai": "Central Mumbai",
    "north-mumbai": "North Mumbai",
}

EXCEL_FILES = [
    "Mumbai96_WestMumbai_SetA.xlsx",
    "Mumbai96_WestMumbai_SetB.xlsx",
    "Mumbai96_WestMumbai_SetC.xlsx",
    "Mumbai96_CentralMumbai.xlsx",
    "Mumbai96_NorthMumbai_Set1.xlsx",
    "24_SouthMumbai.xlsx",
    # Duplicates of above — idempotent, safe to run
    "11_CentralMumbai.xlsx",
    "12_NorthMumbai_Set1.xlsx",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_str(val):
    """Return stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_float(val):
    """Return float or None."""
    if val is None:
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def get_or_create_root():
    """Ensure 'Mumbai' root location exists."""
    mumbai = Location.query.filter_by(slug="mumbai").first()
    if not mumbai:
        mumbai = Location(name="Mumbai", slug="mumbai", is_active=True)
        db.session.add(mumbai)
        db.session.flush()
        print("  🆕 Created root: Mumbai")
    return mumbai


def get_or_create_zone(zone_slug, zone_name, parent_id):
    """Get or create a zone-level location (e.g. west-mumbai) under the root."""
    zone = Location.query.filter_by(slug=zone_slug).first()
    if not zone:
        zone = Location(
            name=zone_name,
            slug=zone_slug,
            parent_id=parent_id,
            is_active=True,
        )
        db.session.add(zone)
        db.session.flush()
        print(f"  🆕 Created zone: {zone_name} ({zone_slug})")
    return zone


def parse_places(row):
    """Parse 3 places → [{name, tag, description}]"""
    places = []
    for i in range(1, 4):
        name = safe_str(row[COL[f"place{i}_name"]])
        tag = safe_str(row[COL[f"place{i}_tag"]])
        desc = safe_str(row[COL[f"place{i}_desc"]])
        if name or desc:
            entry = {}
            if name:
                entry["name"] = name
            if tag:
                entry["tag"] = tag
            if desc:
                entry["description"] = desc
            places.append(entry)
    return places


def parse_food(row):
    """Parse food → [{name, description, food_type[]}]"""
    items = []
    food_tags_str = safe_str(row[COL["food_tags"]])
    food_tags = [t.strip() for t in food_tags_str.split("|") if t.strip()] if food_tags_str else []

    for i in range(1, 5):
        name = safe_str(row[COL[f"food{i}_name"]])
        desc = safe_str(row[COL[f"food{i}_desc"]])
        if name or desc:
            entry = {}
            if name:
                entry["name"] = name
            if desc:
                entry["description"] = desc
            # Only attach food_type to the first food item
            if i == 1 and food_tags:
                entry["food_type"] = food_tags
            items.append(entry)
    return items


def parse_living_style(row):
    """Parse sub-areas → [{name, description}]"""
    items = []
    for i in range(1, 4):
        name = safe_str(row[COL[f"sub{i}_name"]])
        desc = safe_str(row[COL[f"sub{i}_desc"]])
        if name or desc:
            entry = {}
            if name:
                entry["name"] = name
            if desc:
                entry["description"] = desc
            items.append(entry)
    return items


def parse_connectivity(row):
    """Parse commute → [{name, distance, tag}]"""
    items = []
    railway = safe_str(row[COL["railway_station"]])
    if railway:
        items.append({
            "name": railway,
            "tag": "Railway",
            "distance": safe_str(row[COL["time_to_churchgate"]]) or "",
        })

    metro = safe_str(row[COL["metro"]])
    if metro:
        items.append({"name": metro, "tag": "Metro"})

    highway = safe_str(row[COL["highway"]])
    if highway:
        items.append({"name": highway, "tag": "Road"})

    bus = safe_str(row[COL["bus_routes"]])
    if bus:
        items.append({"name": bus, "tag": "Bus"})

    return items


def parse_nearby_locations(row):
    """Parse nearby area slugs → [location_id, ...]"""
    slugs = []
    for key in ["nearby1_slug", "nearby2_slug", "nearby3_slug",
                 "nearby4_slug", "nearby5_slug", "nearby6_slug"]:
        slug = safe_str(row[COL[key]])
        if slug:
            slugs.append(slug.lower())

    ids = []
    for slug in slugs:
        loc = Location.query.filter_by(slug=slug).first()
        if loc:
            ids.append(loc.id)
        else:
            print(f"    ⚠️  Nearby location not found in DB: {slug}")
    return ids


def parse_best_services(row):
    """Parse prose section → [{name, description}]"""
    h2 = safe_str(row[COL["prose_h2"]])
    paras = []
    for c in ["prose_p1", "prose_p2", "prose_p3", "prose_about"]:
        p = safe_str(row[COL[c]])
        if p:
            paras.append(p)

    if h2 or paras:
        items = []
        combined = " ".join(paras) if paras else ""
        if h2:
            items.append({"name": h2, "description": combined})
        elif combined:
            items.append({"name": "Best Services", "description": combined})
        return items
    return []


def parse_sub_areas(row):
    """Parse sub-areas → [{name, tag, description}]"""
    items = []
    for i in range(1, 4):
        name = safe_str(row[COL.get(f"sub{i}_name")])
        tag = safe_str(row[COL.get(f"sub{i}_tag")])
        desc = safe_str(row[COL.get(f"sub{i}_desc")])
        if name or desc:
            entry = {}
            if name:
                entry["name"] = name
            if tag:
                entry["tag"] = tag
            if desc:
                entry["description"] = desc
            items.append(entry)
    return items


def parse_property_prices(row):
    """Parse property prices → [{type, price, sub_area, rental}]"""
    items = []
    for i in range(1, 4):
        ptype = safe_str(row[COL.get(f"prop{i}_type")])
        price = safe_str(row[COL.get(f"prop{i}_price")])
        subarea = safe_str(row[COL.get(f"prop{i}_subarea")])
        if ptype or price:
            entry = {"type": ptype or ""}
            if price:
                entry["price"] = price
            if subarea:
                entry["sub_area"] = subarea
            items.append(entry)
    rental = safe_str(row[COL.get("rental_range")])
    if rental and items:
        # Attach rental to each item, or store separately
        for item in items:
            item["rental"] = rental
    return items


def parse_schools(row):
    """Parse schools → [{name, board}]"""
    items = []
    for i in range(1, 5):
        name = safe_str(row[COL.get(f"school{i}_name")])
        board = safe_str(row[COL.get(f"school{i}_board")])
        if name:
            entry = {"name": name}
            if board:
                entry["board"] = board
            items.append(entry)
    return items


def parse_hospitals(row):
    """Parse hospitals → [{name, type}]"""
    items = []
    for i in range(1, 4):
        name = safe_str(row[COL.get(f"hospital{i}_name")])
        htype = safe_str(row[COL.get(f"hospital{i}_type")])
        if name:
            entry = {"name": name}
            if htype:
                entry["type"] = htype
            items.append(entry)
    return items


def parse_employers(row):
    """Parse major employers → [{name, sector}]"""
    items = []
    for i in range(1, 4):
        name = safe_str(row[COL.get(f"employer{i}_name")])
        sector = safe_str(row[COL.get(f"employer{i}_sector")])
        if name:
            entry = {"name": name}
            if sector:
                entry["sector"] = sector
            items.append(entry)
    return items


def parse_faq(row):
    """Parse FAQ → [{question, answer}]"""
    items = []
    for i in range(1, 6):
        q = safe_str(row[COL.get(f"faq{i}_q")])
        a = safe_str(row[COL.get(f"faq{i}_a")])
        if q and a:
            items.append({"question": q, "answer": a})
    return items


def parse_civic_data(row):
    """Parse civic data → {ward, assembly_constituency, lok_sabha, police_station}"""
    data = {}
    ward = safe_str(row[COL.get("municipal_ward")])
    if ward:
        data["ward"] = ward
    assembly = safe_str(row[COL.get("assembly_constituency")])
    if assembly:
        data["assembly_constituency"] = assembly
    lok_sabha = safe_str(row[COL.get("lok_sabha_constituency")])
    if lok_sabha:
        data["lok_sabha"] = lok_sabha
    # Police station is not in Excel — admin to fill
    return data if data else None


def parse_food_tags(row):
    """Parse food tags as comma-separated string."""
    tags = safe_str(row[COL.get("food_tags")])
    if tags:
        # Excel uses pipe-separated — normalise to comma+space
        return ", ".join(t.strip() for t in tags.split("|") if t.strip())
    return None


# ---------------------------------------------------------------------------
# Editorial field auto-generation
# ---------------------------------------------------------------------------

# Curated editorial overrides for specific locations.
# These are applied AFTER auto-generation, so they replace generic data.
CURATED_EDITORIAL = {
    "dadar-east": {
        "character_vibe": [
            {"keyword": "Maharashtrian Heartland", "description": "Strong Marathi cultural identity with historic roots."},
            {"keyword": "Commercial & Residential Mix", "description": "Busy commercial corridors alongside dense residential chawls and societies."},
            {"keyword": "Working-class Spirit", "description": "Grounded, community-oriented neighbourhood with a vibrant street life."},
        ],
        "resident_profile": [
            {"segment": "Maharashtrian Families", "percentage": 60, "description": "Multi-generational Maharashtrian families in chawls and societies."},
            {"segment": "Working Professionals", "percentage": 25, "description": "Young professionals working in Parel, Lower Parel and BKC."},
            {"segment": "Students", "percentage": 10, "description": "Students attending nearby colleges and coaching classes."},
            {"segment": "Elderly Residents", "percentage": 5, "description": "Long-time residents in older housing stock."},
        ],
        "local_events": [
            {"name": "Ganesh Chaturthi Celebrations", "date": "August-September", "description": "Elaborate Ganesh pandals and community celebrations across Naigaum and TT Circle."},
            {"name": "Dahi Handi (Janmashtami)", "date": "August", "description": "Traditional Dahi Handi events in the chawls of Dadar East."},
            {"name": "Navratri Garba", "date": "September-October", "description": "Community garba events in local society grounds."},
            {"name": "Mumbai Marathon", "date": "January", "description": "Route passes through key Dadar East junctions annually."},
        ],
        "upcoming_projects": [
            {"name": "Dadar TT Circle Redevelopment", "developer": "BMC", "status": "Planning", "description": "Proposed redevelopment of the historic TT circle area to improve traffic flow and pedestrian amenities."},
            {"name": "Coastal Road (Phase 2)", "developer": "BMC / MMRDA", "status": "Ongoing", "description": "Extended coastal road connectivity from Marine Drive to Kandivali, benefiting Dadar East commuters."},
            {"name": "Building Redevelopment Projects", "developer": "Private Developers", "status": "Ongoing", "description": "Multiple old building redevelopment projects in Naigaum and Hindmata areas under new DCPR norms."},
        ],
        "residential_societies": [
            {"name": "Naigaum CHS", "type": "Co-operative Housing Society", "description": "Well-established co-operative housing society in the heart of Naigaum."},
            {"name": "Dadar East Premises Co-op Society", "type": "Co-operative Housing Society", "description": "Large society complex near Dadar station east exit."},
            {"name": "Hindmata Area Apartments", "type": "Mixed Apartments", "description": "Mix of old and redeveloped apartments around Hindmata cinema junction."},
        ],
        "area_report_card": {"safety": 3, "cleanliness": 2, "green": 2, "transit": 5, "overall": 3.2},
    },
    "dadar-west": {
        "character_vibe": [
            {"keyword": "Marathi Cultural Capital", "description": "Home to Shivaji Park, political rallies and Maharashtrian cultural identity."},
            {"keyword": "Prestigious Residential", "description": "One of Mumbai's most desirable addresses with heritage buildings and sea-link access."},
            {"keyword": "Heritage & Modern Blend", "description": "Portuguese-era churches alongside modern high-rises."},
        ],
        "resident_profile": [
            {"segment": "Affluent Families", "percentage": 45, "description": "Established families in premium apartments near Shivaji Park."},
            {"segment": "Professionals", "percentage": 30, "description": "Doctors, lawyers and business professionals."},
            {"segment": "Elderly Residents", "percentage": 15, "description": "Long-standing residents in heritage buildings."},
            {"segment": "Students & Youth", "percentage": 10, "description": "Students at nearby colleges and coaching institutes."},
        ],
        "local_events": [
            {"name": "Shivaji Park Dussehra Melava", "date": "October", "description": "Annual Dussehra rally at Shivaji Park — one of Mumbai's largest political gatherings."},
            {"name": "Ganesh Chaturthi", "date": "August-September", "description": "Grand celebrations at Shivaji Park and surrounding areas."},
            {"name": "Shivaji Park Cricket Coaching", "date": "Year-round", "description": "Famous cricket coaching nets at Shivaji Park — a Mumbai cricketing institution."},
        ],
        "upcoming_projects": [
            {"name": "Bandra-Worli Sea Link Phase 2", "developer": "MSRDC", "status": "Ongoing", "description": "Extended sea link connectivity improving access to Western suburbs."},
            {"name": "Dadar West Redevelopment", "developer": "BMC & Private", "status": "Ongoing", "description": "Multiple redevelopment projects along Ranade Road and surrounding areas."},
        ],
        "residential_societies": [
            {"name": "Shivaji Park Residency", "type": "Premium Apartments", "description": "High-end apartments facing Shivaji Park with sea views."},
            {"name": "Ranade Road Heritage Buildings", "type": "Heritage Apartments", "description": "Pre-war heritage buildings along the commercial strip."},
            {"name": "Portuguese Church Area Societies", "type": "Co-operative Societies", "description": "Well-established societies in the heritage church neighbourhood."},
        ],
        "area_report_card": {"safety": 4, "cleanliness": 3, "green": 3, "transit": 5, "overall": 3.8},
    },
}


EDITORIAL_FIELDS = [
    "character_vibe",
    "resident_profile",
    "local_events",
    "upcoming_projects",
    "residential_societies",
    "area_report_card",
]


def generate_editorial_fields(location):
    """
    Auto-generate editorial fields for a location if they are null.
    Only fills fields that are None — never overwrites existing data.
    """
    name = location.name
    formatted = name.replace("-", " ").title() if name else name

    if location.character_vibe is None:
        location.character_vibe = [
            {"keyword": "Residential & Commercial Hub", "description": f"{formatted} is a bustling neighbourhood with a mix of homes and businesses."},
            {"keyword": "Growing Neighbourhood", "description": f"Ongoing redevelopment and new infrastructure are shaping {formatted}'s future."},
            {"keyword": "Community-Focused", "description": f"Strong local community with markets, temples and social spaces at its core."},
        ]

    if location.resident_profile is None:
        location.resident_profile = [
            {"segment": "Families", "percentage": 50, "description": f"Multi-generational families residing in societies and chawls across {formatted}."},
            {"segment": "Young Professionals", "percentage": 25, "description": f"Professionals working in nearby commercial hubs from {formatted}."},
            {"segment": "Students", "percentage": 15, "description": f"Students attending local colleges and coaching centres in and around {formatted}."},
            {"segment": "Elderly Residents", "percentage": 10, "description": f"Long-time residents in established housing across {formatted}."},
        ]

    if location.local_events is None:
        location.local_events = [
            {"name": "Ganesh Chaturthi", "date": "Aug-Sep", "description": f"Community Ganesh pandals and celebrations across {formatted}."},
            {"name": "Diwali Celebrations", "date": "Oct-Nov", "description": f"Lighting and festivities in societies and streets of {formatted}."},
            {"name": "Navratri Garba", "date": "Sep-Oct", "description": f"Community garba events organised by local societies in {formatted}."},
            {"name": "Mumbai Marathon Route", "date": "January", "description": f"The annual Mumbai Marathon passes through or near {formatted}."},
        ]

    if location.upcoming_projects is None:
        location.upcoming_projects = [
            {"name": "Metro Connectivity", "developer": "MMRDA", "status": "Ongoing", "description": f"Upcoming metro lines improving connectivity for {formatted} residents."},
            {"name": "Building Redevelopment", "developer": "Private Developers", "status": "Ongoing", "description": f"Multiple old buildings in {formatted} are being redeveloped under new DCPR norms."},
            {"name": "Road & Infrastructure Upgrades", "developer": "BMC", "status": "Planning", "description": f"BMC has proposed road widening and drainage improvements for {formatted}."},
        ]

    if location.residential_societies is None:
        location.residential_societies = [
            {"name": f"{formatted} Co-operative Housing Society", "type": "Co-operative Society", "description": f"Well-established co-operative housing society in {formatted}."},
            {"name": f"{formatted} Premises CHS", "type": "Co-operative Society", "description": f"Large residential society complex in the heart of {formatted}."},
            {"name": f"{formatted} Apartments", "type": "Mixed Apartments", "description": "Mix of older and newly redeveloped apartment buildings."},
        ]

    if location.area_report_card is None:
        location.area_report_card = {"safety": 3, "cleanliness": 2, "green": 2, "transit": 4, "overall": 3.0}

    # ── Apply curated overrides (replace whatever was set above) ──
    curated = CURATED_EDITORIAL.get(location.slug)
    if curated:
        for field in EDITORIAL_FIELDS:
            if field in curated:
                setattr(location, field, curated[field])


# Slug → filename overrides for icons that don't match the slug naming convention.
SLUG_TO_ICON_FILENAME = {
    "nalasopara-east": "nala-sopara-east.png",
    "nalasopara-west": "nala-sopara-west.png",
}


def upload_icon(app, slug):
    """Copy icon from data/location-icon/ → static/uploads/location/

    Looks for {slug}.png by default; checks SLUG_TO_ICON_FILENAME for
    mismatched filenames.

    Returns the relative path for the DB column, or None.
    """
    icon_path = None

    # 1. Check explicit mapping first
    mapped = SLUG_TO_ICON_FILENAME.get(slug)
    if mapped:
        candidate = os.path.join(ICON_DIR, mapped)
        if os.path.isfile(candidate):
            icon_path = candidate

    # 2. Fallback to {slug}.{ext}
    if not icon_path:
        for ext in [".png", ".jpg", ".jpeg", ".webp"]:
            candidate = os.path.join(ICON_DIR, f"{slug}{ext}")
            if os.path.isfile(candidate):
                icon_path = candidate
                break

    if not icon_path:
        return None

    ext = os.path.splitext(icon_path)[1].lower()
    unique_name = f"{uuid.uuid4().hex}{ext}"
    rel_path = f"location/{unique_name}"

    upload_folder = app.config.get("UPLOAD_FOLDER", "static/uploads")
    # app.root_path = backend/app/
    full_dir = os.path.normpath(os.path.join(app.root_path, "..", upload_folder, "location"))
    full_path = os.path.join(full_dir, unique_name)
    os.makedirs(full_dir, exist_ok=True)
    shutil.copy2(icon_path, full_path)

    return rel_path


# ---------------------------------------------------------------------------
# Cleanup helpers
# ---------------------------------------------------------------------------

LEGACY_SLUGS_TO_DELETE = [
    "bhandup-east", "bhandup-west",
    "kurla-east", "kurla-west",
    "vikhroli-east", "vikhroli-west",
    "mahim-east", "mahim-west",
    "matunga-east", "matunga-west",
    "grant-road-east", "grant-road-west",
    "kanjurmarg-east", "kanjurmarg-west",
    "andheri-east",
]


def delete_location_and_refs(location, deleted_ids=None):
    """Delete a location and all FK references pointing to it."""
    if not location:
        return

    deleted_ids = deleted_ids if deleted_ids is not None else set()
    LocationCategory.query.filter_by(location_id=location.id).delete()
    Listing.query.filter_by(location_id=location.id).delete()
    SeoAliasSlug.query.filter(
        (SeoAliasSlug.location_id == location.id) |
        (SeoAliasSlug.canonical_location_id == location.id)
    ).delete()

    for loc in Location.query.filter(Location.nearby_locations.isnot(None)).all():
        if not loc.nearby_locations:
            continue
        cleaned = [lid for lid in loc.nearby_locations if lid != location.id]
        if len(cleaned) != len(loc.nearby_locations):
            loc.nearby_locations = cleaned if cleaned else None

    db.session.flush()
    db.session.delete(location)
    deleted_ids.add(location.id)
    return deleted_ids


def collect_expected_location_slugs_from_excel():
    """Return the exact valid slugs for each zone from the Excel source files."""
    expected = {zone_slug: set() for zone_slug in ZONE_MAP.values()}

    for fname in EXCEL_FILES:
        filepath = os.path.join(DATA_DIR, fname)
        if not os.path.isfile(filepath):
            continue

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(min_row=3, values_only=True):
            slug = safe_str(row[COL["slug"]])
            if not slug:
                continue
            slug = slug.lower()
            set_val = safe_str(row[COL["set"]])
            zone_slug = ZONE_MAP.get(set_val)
            if zone_slug:
                expected.setdefault(zone_slug, set()).add(slug)

        wb.close()

    return expected


def repair_zone_locations():
    """Strip stale or misassigned rows so each zone keeps only the exact Excel-backed locations."""
    expected_by_zone = collect_expected_location_slugs_from_excel()
    deleted_ids = set()
    stats = {"deleted": 0, "reparented": 0}

    for zone_slug, expected_slugs in expected_by_zone.items():
        zone = Location.query.filter_by(slug=zone_slug).first()
        if not zone:
            continue

        for loc in Location.query.filter(Location.parent_id == zone.id, Location.slug != zone_slug).all():
            if loc.slug.lower() not in expected_slugs:
                delete_location_and_refs(loc, deleted_ids)
                stats["deleted"] += 1
                print(f"  🧹 Removed stale row: {loc.slug} (zone={zone_slug})")

        for loc in Location.query.filter(Location.slug.in_(sorted(expected_slugs))).all():
            if loc.parent_id != zone.id:
                loc.parent_id = zone.id
                stats["reparented"] += 1
                print(f"  🔗 Reparented exact row: {loc.slug} → {zone_slug}")

    if deleted_ids:
        for loc in Location.query.filter(Location.nearby_locations.isnot(None)).all():
            if not loc.nearby_locations:
                continue
            cleaned = [lid for lid in loc.nearby_locations if lid not in deleted_ids]
            if len(cleaned) != len(loc.nearby_locations):
                loc.nearby_locations = cleaned if cleaned else None

    db.session.flush()
    return stats


def clean_legacy_hierarchy():
    """Remove old/duplicate entries and fix up the location hierarchy.

    Steps:
        1. Delete western-mumbai (replaced by west-mumbai)
        2. Delete 15 legacy east/west locations under capital-M Mumbai
        3. Delete capital-M Mumbai itself
        4. Reparent south-mumbai under lowercase mumbai
        5. Remove dangling nearby_locations references to deleted IDs
    """
    deleted_ids = set()
    stats = {"deleted": 0, "reparented": 0, "refs_cleaned": 0}

    # ── 1. Delete western-mumbai ──
    wm = Location.query.filter_by(slug="western-mumbai").first()
    if wm:
        delete_location_and_refs(wm, deleted_ids)
        stats["deleted"] += 1
        print(f"  🗑️  Deleted: western-mumbai (id={wm.id})")

    # ── 2. Delete 15 legacy east/west locations ──
    for loc in Location.query.filter(Location.slug.in_(LEGACY_SLUGS_TO_DELETE)).all():
        delete_location_and_refs(loc, deleted_ids)
        stats["deleted"] += 1
        print(f"  🗑️  Deleted: {loc.slug} (id={loc.id})")

    # ── 3. Delete capital-M Mumbai (safety: only if 0 children remain) ──
    mumbai_cap = Location.query.filter_by(slug="Mumbai").first()
    if mumbai_cap:
        remaining = Location.query.filter_by(parent_id=mumbai_cap.id).count()
        if remaining == 0:
            delete_location_and_refs(mumbai_cap, deleted_ids)
            stats["deleted"] += 1
            print(f"  🗑️  Deleted: Mumbai (capital-M, id={mumbai_cap.id})")
        else:
            print(f"  ⚠️  Capital-M Mumbai still has {remaining} children — skipping")

    # ── 4. Reparent south-mumbai under lowercase mumbai ──
    root = Location.query.filter_by(slug="mumbai").first()
    south = Location.query.filter_by(slug="south-mumbai").first()
    if south and root and south.parent_id != root.id:
        south.parent_id = root.id
        stats["reparented"] += 1
        print(f"  🔗 Reparented: south-mumbai (id={south.id}) → mumbai (id={root.id})")

    # ── 5. Remove dangling nearby_locations references ──
    if deleted_ids:
        for loc in Location.query.filter(Location.nearby_locations.isnot(None)).all():
            if not loc.nearby_locations:
                continue
            cleaned = [lid for lid in loc.nearby_locations if lid not in deleted_ids]
            if len(cleaned) != len(loc.nearby_locations):
                loc.nearby_locations = cleaned if cleaned else None
                stats["refs_cleaned"] += 1

    db.session.flush()
    return stats


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def process_excel(app, filepath):
    """Process a single Excel file, returning a stats dict."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    stats = {"created": 0, "updated": 0, "skipped": 0, "icons": 0, "editorial_seeded": 0}

    for row in ws.iter_rows(min_row=3, values_only=True):
        slug = safe_str(row[COL["slug"]])
        if not slug:
            continue
        slug = slug.lower()
        name = safe_str(row[COL["name"]]) or slug.replace("-", " ").title()

        # --- Zone / parent ---
        set_val = safe_str(row[COL["set"]])
        zone_slug = ZONE_MAP.get(set_val)
        if not zone_slug:
            print(f"  ⚠️  Unknown set '{set_val}' for {slug}, skipping")
            stats["skipped"] += 1
            continue

        zone = Location.query.filter_by(slug=zone_slug).first()
        if not zone:
            print(f"  ⚠️  Zone '{zone_slug}' not found for {slug}, skipping")
            stats["skipped"] += 1
            continue

        # --- Find or create location ---
        location = Location.query.filter_by(slug=slug).first()
        is_new = False
        if not location:
            location = Location(name=name, slug=slug, parent_id=zone.id, is_active=True)
            db.session.add(location)
            db.session.flush()
            is_new = True
            print(f"  🆕 Created: {name} ({slug})")

        # --- Basic fields ---
        location.name = name
        location.parent_id = zone.id  # ensure parent is set
        location.is_active = True  # ensure imported rows stay visible to public APIs

        lat = safe_float(row[COL["lat"]])
        if lat is not None:
            location.lat = lat
        lng = safe_float(row[COL["lng"]])
        if lng is not None:
            location.lng = lng

        if seo_title := safe_str(row[COL["seo_title"]]):
            location.seo_title = seo_title
        if seo_desc := safe_str(row[COL["seo_description"]]):
            location.seo_description = seo_desc
        if seo_kw := safe_str(row[COL["seo_keywords"]]):
            location.seo_keywords = seo_kw
        if population := safe_str(row[COL["population"]]):
            location.population = population
        if municipal_body := safe_str(row[COL["municipal_body"]]):
            location.municipal_body = municipal_body

        # --- About (combine para1 + para2) ---
        about_p1 = safe_str(row[COL["about_p1"]])
        about_p2 = safe_str(row[COL["about_p2"]])
        if about_p1 or about_p2:
            location.about = "\n\n".join(p for p in [about_p1, about_p2] if p)

        # --- JSON fields ---
        try:
            places = parse_places(row)
            if places:
                location.places_to_visit = places

            food = parse_food(row)
            if food:
                location.food = food

            # Sub-areas (with tags) — stored in sub_areas AND living_style for backward compat
            sub_areas = parse_sub_areas(row)
            if sub_areas:
                location.sub_areas = sub_areas
                # Also keep living_style populated for backward compat (omit tag)
                location.living_style = [
                    {k: v for k, v in sa.items() if k != "tag"}
                    for sa in sub_areas
                ]

            # Property prices
            property_prices = parse_property_prices(row)
            if property_prices:
                location.property_prices = property_prices

            # Schools
            schools = parse_schools(row)
            if schools:
                location.schools = schools

            # Hospitals
            hospitals = parse_hospitals(row)
            if hospitals:
                location.hospitals = hospitals

            # Major employers
            employers = parse_employers(row)
            if employers:
                location.major_employers = employers

            # FAQ
            faq = parse_faq(row)
            if faq:
                location.faq = faq

            # Civic data
            civic = parse_civic_data(row)
            if civic:
                location.civic_data = civic

            # Food tags (separate from food items)
            food_tags = parse_food_tags(row)
            if food_tags:
                location.food_tags = food_tags

            connectivity = parse_connectivity(row)
            if connectivity:
                location.travelling_connectivity = connectivity

            services = parse_best_services(row)
            if services:
                location.best_services = services

            # night_life – not in Excel, leave as empty array
            location.night_life = []

            nearby_ids = parse_nearby_locations(row)
            if nearby_ids:
                location.nearby_locations = nearby_ids
        except Exception as e:
            print(f"  ❌ Error parsing JSON fields for {slug}: {e}")
            stats["skipped"] += 1
            continue

        # --- Icon ---
        icon_rel = upload_icon(app, slug)
        if icon_rel:
            location.location_icon = icon_rel
            stats["icons"] += 1

        # --- Editorial fields (auto-generate if null) ---
        generate_editorial_fields(location)

        db.session.flush()

        if is_new:
            stats["created"] += 1
        else:
            stats["updated"] += 1

    wb.close()
    return stats


def main():
    print("=" * 60)
    print("  Mumbai96 – Location Metadata Import")
    print("=" * 60)

    app = create_app()

    with app.app_context():
        # ── Ensure root + zone parents ──
        root = get_or_create_root()
        db.session.commit()

        for zone_slug, zone_name in sorted(ZONE_DISPLAY.items()):
            get_or_create_zone(zone_slug, zone_name, root.id)
        db.session.commit()

        print(f"\n  Root: Mumbai (id={root.id})")
        for zs in sorted(ZONE_DISPLAY):
            z = Location.query.filter_by(slug=zs).first()
            if z:
                print(f"  Zone: {z.name:20s} (id={z.id:3d}, slug={zs})")

        print("\n" + "-" * 60)
        print("  Cleaning stale exact-area rows …")
        zone_stats = repair_zone_locations()
        db.session.commit()
        print(f"    Removed stale rows: {zone_stats['deleted']}")
        print(f"    Reparented exact rows: {zone_stats['reparented']}")

        print("\n" + "-" * 60)

        # ── Process Excel files ──
        grand = {"created": 0, "updated": 0, "skipped": 0, "icons": 0}

        for fname in EXCEL_FILES:
            filepath = os.path.join(DATA_DIR, fname)
            if not os.path.isfile(filepath):
                print(f"\n  ⏭️  {fname} – file not found")
                continue

            print(f"\n  📄 {fname} …")
            stats = process_excel(app, filepath)
            db.session.commit()

            for k in grand:
                grand[k] += stats[k]
            print(f"     → {stats['created']} created, {stats['updated']} updated, "
                  f"{stats['skipped']} skipped, {stats['icons']} icons")

        # ── Editorial seed pass for ALL active locations ──
        # This ensures zones and any location not in Excel also get editorial fields.
        print("\n" + "-" * 60)
        print("  Seeding editorial fields …")
        all_locations = Location.query.filter_by(is_active=True).all()
        editorial_seeded = 0
        editorial_skipped = 0
        for loc in all_locations:
            # Check if any editorial field is None
            needs_seed = any(
                getattr(loc, field) is None
                for field in EDITORIAL_FIELDS
            )
            if needs_seed:
                generate_editorial_fields(loc)
                db.session.flush()
                editorial_seeded += 1
            else:
                editorial_skipped += 1
        db.session.commit()
        print(f"    Seeded: {editorial_seeded}, Already had data: {editorial_skipped}")

        # ── Clean up legacy hierarchy ──
        print("\n" + "-" * 60)
        print("  Cleaning legacy hierarchy …")
        clean_stats = clean_legacy_hierarchy()
        db.session.commit()
        print(f"    Deleted:            {clean_stats['deleted']}")
        print(f"    Reparented:         {clean_stats['reparented']}")
        print(f"    Nearby refs fixed:  {clean_stats['refs_cleaned']}")

        # ── Summary ──
        print("\n" + "=" * 60)
        print("  SUMMARY")
        print(f"    Locations created:  {grand['created']}")
        print(f"    Locations updated:  {grand['updated']}")
        print(f"    Skipped:            {grand['skipped']}")
        print(f"    Icons imported:     {grand['icons']}")
        print(f"    Editorial seeded:   {editorial_seeded}")
        print(f"    Legacy deleted:     {clean_stats['deleted']}")
        print(f"    Reparented:         {clean_stats['reparented']}")
        print(f"    Nearby refs fixed:  {clean_stats['refs_cleaned']}")
        print("=" * 60)


if __name__ == "__main__":
    main()
