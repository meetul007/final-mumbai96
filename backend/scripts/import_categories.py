"""
Import categories with emojis and Google Places API types from Excel.

Also creates LocationCategory records for any new categories,
linking them to every active location.

Usage:
    cd backend
    python scripts/import_categories.py
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from slugify import slugify
from app import create_app
from app.extensions import db
from app.listing.models import Category, Location, LocationCategory

import openpyxl

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
EXCEL_PATH = os.path.join(DATA_DIR, "category", "186_categories_emojis.xlsx")

# Categories that exist in DB but NOT in the Excel — these are old/typo entries
# that should be marked as inactive rather than deleted (to preserve FK refs).
DEPRECATED_SLUGS = [
    "accountant",
    "accounting-firm",
    "auditor",
    "certified-public-accountant",
    "chartered-accountant",
    "coaching-center",
    "consultant",
    "corporate-office",
    "educational-institution",
    "gyms-",
    "personal-trainers-",
    "real-estate-attorney",
    "relationship-councellor",
    "sports-complex",
    "tatoo-artists",
    "tax-consultant",
]


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def run():
    app = create_app()

    with app.app_context():
        print("=" * 60)
        print("  Importing categories with emojis")
        print("=" * 60)

        if not os.path.isfile(EXCEL_PATH):
            print(f"  ❌ File not found: {EXCEL_PATH}")
            return

        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        ws = wb[wb.sheetnames[0]]

        stats = {"created": 0, "updated": 0, "lc_created": 0, "deprecated": 0}

        # ── Get all active locations (for LocationCategory creation) ──
        all_locations = Location.query.filter_by(is_active=True).all()
        location_count = len(all_locations)

        # ── Process each Excel row ──
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_id = row[0]           # #
            slug = safe_str(row[1])   # Category slug
            emoji = safe_str(row[2])  # Emoji
            gplace = safe_str(row[3]) # Google Places API Type

            if not slug:
                continue

            slug = slug.lower()
            name = slug.replace("-", " ").title()

            category = Category.query.filter_by(slug=slug).first()
            is_new = False

            if category:
                # ── Update existing ──
                if emoji and category.emoji != emoji:
                    category.emoji = emoji
                if gplace:
                    category.google_places_api_type = gplace
                if not category.name:
                    category.name = name
                stats["updated"] += 1

                # ── Backfill missing LocationCategory records ──
                existing_loc_ids = set(
                    row[0] for row in db.session.query(LocationCategory.location_id)
                    .filter(LocationCategory.category_id == category.id)
                    .all()
                )
                for loc in all_locations:
                    if loc.id not in existing_loc_ids:
                        base_slug = slugify(f"{category.name or name} {loc.name}")
                        lc_slug = f"{base_slug}-{category.id}"
                        lc = LocationCategory(
                            location_id=loc.id,
                            category_id=category.id,
                            name=f"{category.name or name} in {loc.name}",
                            description=f"{category.name or name} in {loc.name}",
                            seo_description=f"{category.name or name} in {loc.name}",
                            seo_slug=lc_slug,
                        )
                        db.session.add(lc)
                        stats["lc_created"] += 1

                if stats["lc_created"] > 0:
                    db.session.flush()
            else:
                # ── Create new ──
                category = Category(
                    name=name,
                    slug=slug,
                    emoji=emoji,
                    google_places_api_type=gplace,
                    is_active=True,
                )
                db.session.add(category)
                db.session.flush()
                is_new = True
                stats["created"] += 1
                print(f"  🆕 Created category: {name} ({slug}) {emoji or ''}")

                # ── Create LocationCategory for ALL locations ──
                for loc in all_locations:
                    existing_lc = LocationCategory.query.filter_by(
                        location_id=loc.id, category_id=category.id
                    ).first()
                    if not existing_lc:
                        # Make seo_slug unique by appending category id to avoid
                        # conflicts with old categories (e.g. "gyms-" vs "gyms")
                        base_slug = slugify(f"{name} {loc.name}")
                        lc_slug = f"{base_slug}-{category.id}"
                        lc = LocationCategory(
                            location_id=loc.id,
                            category_id=category.id,
                            name=f"{name} in {loc.name}",
                            description=f"{name} in {loc.name}",
                            seo_description=f"{name} in {loc.name}",
                            seo_slug=lc_slug,
                        )
                        db.session.add(lc)
                        stats["lc_created"] += 1

                if stats["lc_created"] > 0:
                    db.session.flush()

            # For existing categories, also ensure emoji is set
            if emoji and not category.emoji:
                category.emoji = emoji

        wb.close()

        # ── Mark deprecated categories as inactive ──
        for old_slug in DEPRECATED_SLUGS:
            cat = Category.query.filter_by(slug=old_slug).first()
            if cat and cat.is_active:
                cat.is_active = False
                stats["deprecated"] += 1
                print(f"  🚫 Deactivated deprecated: {cat.name} ({cat.slug})")

        db.session.commit()

        # ── Summary ──
        print("\n" + "=" * 60)
        print("  SUMMARY")
        print(f"    Categories created:     {stats['created']}")
        print(f"    Categories updated:     {stats['updated']}")
        print(f"    LocationCategory added: {stats['lc_created']}")
        print(f"    Deprecated (inactive):  {stats['deprecated']}")
        print("=" * 60)


if __name__ == "__main__":
    run()
